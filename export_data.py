#!/usr/bin/env python
import os
import sys
import django
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import defaultdict

# Setup Django environment
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ycs_mp.settings')
django.setup()

from events.models import EventRegistration, UpZone, ResponsibilityOption

def find_duplicates():
    """Find duplicate registrations using same logic as duplicate_remove.py"""
    duplicate_groups = defaultdict(list)
    
    for registration in EventRegistration.objects.all().order_by('registration_date'):
        key = (
            registration.gender.lower() if registration.gender else '',
            registration.full_name.lower().strip(),
            registration.date_of_birth,
            registration.state.lower().strip() if registration.state else '',
            registration.city.lower().strip() if registration.city else ''
        )
        duplicate_groups[key].append(registration)
    
    duplicates = {key: regs for key, regs in duplicate_groups.items() if len(regs) > 1}
    return duplicates

# Cache UpZones to avoid repeated queries
UPZONE_CACHE = {}

def get_upzone_cache():
    """Build UpZone cache once"""
    global UPZONE_CACHE
    if not UPZONE_CACHE:
        for upzone in UpZone.objects.filter(is_active=True):
            for district in upzone.districts or []:
                UPZONE_CACHE[district] = upzone.name
    return UPZONE_CACHE

def get_upzone_for_district(city, state):
    """Get UpZone name for a district using cache"""
    if state and 'madhya pradesh' in state.lower():
        cache = get_upzone_cache()
        return cache.get(city, 'No UpZone')
    return 'Not MP'

def get_document_info(registration):
    """Get all document URLs as clickable links"""
    base_url = "https://ycsmp.in"
    
    raw_docs = {
        'aadhar_full': getattr(registration, 'aadhar_full', '') or '',
        'aadhar_front': getattr(registration, 'aadhar_front', '') or '',
        'aadhar_back': getattr(registration, 'aadhar_back', '') or '',
        'passport_photo': getattr(registration, 'passport_photo', '') or '',
        'aadhar_upload_type': getattr(registration, 'aadhar_upload_type', '') or ''
    }
    
    # Convert to full URLs
    docs = {}
    for key, url in raw_docs.items():
        if key == 'aadhar_upload_type':
            docs[key] = url
        elif url:
            docs[key] = f"{base_url}{url}" if not url.startswith('http') else url
        else:
            docs[key] = ''
    
    # Status checks
    has_aadhar_full = bool(raw_docs['aadhar_full'])
    has_aadhar_front = bool(raw_docs['aadhar_front'])
    has_aadhar_back = bool(raw_docs['aadhar_back'])
    has_passport = bool(raw_docs['passport_photo'])
    
    aadhar_status = 'None'
    if has_aadhar_full:
        aadhar_status = 'Full Aadhar'
    elif has_aadhar_front and has_aadhar_back:
        aadhar_status = 'Front & Back'
    elif has_aadhar_front or has_aadhar_back:
        aadhar_status = 'Partial'
    
    return docs, aadhar_status, has_passport

def export_responsibility_options():
    """Export responsibility options data to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Responsibility Options"
    
    headers = ['ID', 'Name', 'Order', 'Is Active', 'Created At']
    ws.append(headers)
    
    for option in ResponsibilityOption.objects.all().order_by('order', 'name'):
        ws.append([
            option.id,
            option.name,
            option.order,
            option.is_active,
            option.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    filename = "responsibility_options.xlsx"
    wb.save(filename)
    print(f"Exported responsibility options to {filename}")
    return filename

def export_to_excel():
    """Export data with unique and duplicate records in separate sheets"""
    
    # Count unique records first
    duplicates = find_duplicates()
    duplicate_ids = set()
    for regs in duplicates.values():
        duplicate_ids.update(reg.id for reg in regs)
    
    total_records = EventRegistration.objects.count()
    unique_count_preview = total_records - len(duplicate_ids)
    
    filename = f"event_registrations_{unique_count_preview}.xlsx"
    wb = Workbook()
    
    headers = [
        'ID', 'Registration Number', 'Full Name', 'Email', 'Phone', 'Gender', 'Date of Birth', 
        'Education', 'Occupation', 'Village/Taluka', 'City', 'State', 'UpZone', 'Registration Type', 
        'Event Title', 'Registration Date', 'Approval Status', 'Vehicle', 'Vehicle Number',
        'Arrival Date', 'Volunteering Details', 'Selected Campaigns', 'Selected Vibhags', 'Responsibility',
        'Aadhar Upload Type', 'Aadhar Full URL', 'Aadhar Front URL', 
        'Aadhar Back URL', 'Passport Photo URL', 'Aadhar Status', 'Has Passport Photo'
    ]
    
    # Use already found duplicates
    duplicate_ids = set()
    for regs in duplicates.values():
        duplicate_ids.update(reg.id for reg in regs)
    
    # Unique records sheet
    ws_unique = wb.active
    ws_unique.title = "Unique Records"
    ws_unique.append(headers)
    
    # Batch process unique records
    unique_regs = EventRegistration.objects.select_related('event').exclude(id__in=duplicate_ids)
    unique_data = []
    
    for reg in unique_regs:
        docs, aadhar_status, has_passport = get_document_info(reg)
        upzone = get_upzone_for_district(reg.city, reg.state)
        row_data = [
            reg.id, getattr(reg, 'registration_number', ''), reg.full_name, 
            getattr(reg, 'email', ''), getattr(reg, 'phone', ''), reg.gender, reg.date_of_birth,
            getattr(reg, 'education', ''), getattr(reg, 'occupation', ''), 
            getattr(reg, 'village_taluka', ''), reg.city, reg.state, upzone, reg.get_registration_type_display(),
            reg.event.title, reg.registration_date.strftime('%Y-%m-%d %H:%M:%S'),
            reg.get_approval_status_display(), getattr(reg, 'transport_mode', ''), 
            getattr(reg, 'vehicle_number', ''), getattr(reg, 'arrival_date', ''),
            getattr(reg, 'volunteering_details', ''), reg.get_campaign_names(), reg.get_vibhag_names(),
            reg.responsibility.name if reg.responsibility else '',
            docs['aadhar_upload_type'], docs['aadhar_full'], docs['aadhar_front'],
            docs['aadhar_back'], docs['passport_photo'], aadhar_status, has_passport
        ]
        unique_data.append((row_data, docs))
    
    # Write all unique data at once
    for row_data, docs in unique_data:
        ws_unique.append(row_data)
    
    unique_count = len(unique_data)
    
    # Duplicates sheet
    ws_dup = wb.create_sheet("Duplicate Records")
    ws_dup.append(headers + ['Duplicate Group'])
    
    # Batch process duplicate records
    duplicate_data = []
    for i, (key, regs) in enumerate(duplicates.items(), 1):
        for reg in regs:
            docs, aadhar_status, has_passport = get_document_info(reg)
            upzone = get_upzone_for_district(reg.city, reg.state)
            row_data = [
                reg.id, getattr(reg, 'registration_number', ''), reg.full_name,
                getattr(reg, 'email', ''), getattr(reg, 'phone', ''), reg.gender, reg.date_of_birth,
                getattr(reg, 'education', ''), getattr(reg, 'occupation', ''),
                getattr(reg, 'village_taluka', ''), reg.city, reg.state, upzone, reg.get_registration_type_display(),
                reg.event.title, reg.registration_date.strftime('%Y-%m-%d %H:%M:%S'),
                reg.get_approval_status_display(), getattr(reg, 'transport_mode', ''),
                getattr(reg, 'vehicle_number', ''), getattr(reg, 'arrival_date', ''),
                getattr(reg, 'volunteering_details', ''), reg.get_campaign_names(), reg.get_vibhag_names(),
                reg.responsibility.name if reg.responsibility else '',
                docs['aadhar_upload_type'], docs['aadhar_full'], docs['aadhar_front'],
                docs['aadhar_back'], docs['passport_photo'], aadhar_status, has_passport, f"Group {i}"
            ]
            duplicate_data.append((row_data, docs))
    
    # Write all duplicate data at once
    for row_data, docs in duplicate_data:
        ws_dup.append(row_data)
    
    duplicate_count = len(duplicate_data)
    
    wb.save(filename)
    print(f"Exported {unique_count} unique records and {duplicate_count} duplicate records to {filename}")
    return filename

if __name__ == "__main__":
    export_to_excel()
    export_responsibility_options()