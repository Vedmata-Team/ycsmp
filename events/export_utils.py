import csv
import io
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

class ExportManager:
    @staticmethod
    def export_to_csv(queryset, filename, fields):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        response.write('\ufeff')  # BOM for UTF-8
        
        writer = csv.writer(response)
        headers = [field['header'] for field in fields]
        writer.writerow(headers)
        
        for obj in queryset:
            row = []
            for field in fields:
                value = ExportManager._get_field_value(obj, field['field'])
                row.append(str(value) if value is not None else '')
            writer.writerow(row)
        
        return response
    
    @staticmethod
    def export_to_excel(queryset, filename, fields):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Export"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = [field['header'] for field in fields]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = ExportManager._get_field_value(obj, field['field'])
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def export_to_pdf(queryset, filename, fields, title="Data Export"):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        story = []
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # Prepare table data
        headers = [field['header'] for field in fields]
        data = [headers]
        
        for obj in queryset:
            row = []
            for field in fields:
                value = ExportManager._get_field_value(obj, field['field'])
                row.append(str(value) if value is not None else '')
            data.append(row)
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response
    
    @staticmethod
    def _get_field_value(obj, field_path):
        """Get nested field value from object"""
        try:
            value = obj
            for attr in field_path.split('.'):
                if hasattr(value, attr):
                    value = getattr(value, attr)
                    if callable(value):
                        value = value()
                else:
                    return None
            
            # Format specific data types
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, list):
                return ', '.join(str(v) for v in value)
            elif hasattr(value, '__str__'):
                return str(value)
            
            return value
        except:
            return None

# Field definitions for different models
EVENT_FIELDS = [
    {'field': 'title', 'header': 'कार्यक्रम नाम'},
    {'field': 'category', 'header': 'श्रेणी'},
    {'field': 'venue', 'header': 'स्थान'},
    {'field': 'district', 'header': 'जिला'},
    {'field': 'event_date', 'header': 'कार्यक्रम तिथि'},
    {'field': 'registration_deadline', 'header': 'पंजीकरण अंतिम तिथि'},
    {'field': 'registration_fee', 'header': 'पंजीकरण शुल्क'},
    {'field': 'max_participants', 'header': 'अधिकतम प्रतिभागी'},
    {'field': 'registered_count', 'header': 'पंजीकृत संख्या'},
    {'field': 'is_published', 'header': 'प्रकाशित'},
    {'field': 'created_at', 'header': 'बनाया गया'}
]

REGISTRATION_FIELDS = [
    {'field': 'registration_number', 'header': 'पंजीकरण संख्या'},
    {'field': 'full_name', 'header': 'नाम'},
    {'field': 'email', 'header': 'ईमेल'},
    {'field': 'phone', 'header': 'मोबाइल'},
    {'field': 'date_of_birth', 'header': 'जन्म तिथि'},
    {'field': 'gender', 'header': 'लिंग'},
    {'field': 'education', 'header': 'शिक्षा'},
    {'field': 'occupation', 'header': 'व्यवसाय'},
    {'field': 'village_taluka', 'header': 'गांव/तालुका'},
    {'field': 'city', 'header': 'जिला'},
    {'field': 'state', 'header': 'राज्य'},
    {'field': 'event.title', 'header': 'कार्यक्रम'},
    {'field': 'approval_status', 'header': 'अप्रूवल स्थिति'},
    {'field': 'registration_date', 'header': 'पंजीकरण तिथि'},
    {'field': 'bringing_vehicle', 'header': 'वाहन'},
    {'field': 'vehicle_number', 'header': 'वाहन नंबर'},
    {'field': 'arrival_date', 'header': 'आगमन तिथि'},
    {'field': 'departure_date', 'header': 'प्रस्थान तिथि'}
]

APPROVAL_USER_FIELDS = [
    {'field': 'user.username', 'header': 'यूजरनेम'},
    {'field': 'user.first_name', 'header': 'नाम'},
    {'field': 'user.email', 'header': 'ईमेल'},
    {'field': 'state_code', 'header': 'राज्य कोड'},
    {'field': 'is_state_approver', 'header': 'राज्य अप्रूवर'},
    {'field': 'is_district_approver', 'header': 'जिला अप्रूवर'},
    {'field': 'districts', 'header': 'असाइन जिले'},
    {'field': 'get_assignment_display', 'header': 'असाइनमेंट'}
]