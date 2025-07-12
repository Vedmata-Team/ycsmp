from django.core.management.base import BaseCommand
from django.core.management import CommandError
from events.models import Event, EventRegistration, ApprovalUser
from events.export_utils import ExportManager, EVENT_FIELDS, REGISTRATION_FIELDS, APPROVAL_USER_FIELDS
import os

class Command(BaseCommand):
    help = 'Export data to CSV, Excel, or PDF files'

    def add_arguments(self, parser):
        parser.add_argument(
            '--model',
            type=str,
            choices=['events', 'registrations', 'approval_users', 'all'],
            required=True,
            help='Model to export'
        )
        parser.add_argument(
            '--format',
            type=str,
            choices=['csv', 'excel', 'pdf'],
            default='csv',
            help='Export format'
        )
        parser.add_argument(
            '--output-dir',
            type=str,
            default='exports',
            help='Output directory for exported files'
        )
        parser.add_argument(
            '--event-id',
            type=int,
            help='Export registrations for specific event ID'
        )
        parser.add_argument(
            '--status',
            type=str,
            choices=['pending', 'level1_approved', 'approved', 'rejected'],
            help='Filter registrations by approval status'
        )

    def handle(self, *args, **options):
        model = options['model']
        format_type = options['format']
        output_dir = options['output_dir']
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        if model == 'all':
            self.export_all(format_type, output_dir, options)
        elif model == 'events':
            self.export_events(format_type, output_dir)
        elif model == 'registrations':
            self.export_registrations(format_type, output_dir, options)
        elif model == 'approval_users':
            self.export_approval_users(format_type, output_dir)

    def export_all(self, format_type, output_dir, options):
        self.stdout.write('Exporting all data...')
        self.export_events(format_type, output_dir)
        self.export_registrations(format_type, output_dir, options)
        self.export_approval_users(format_type, output_dir)
        self.stdout.write(self.style.SUCCESS('All exports completed!'))

    def export_events(self, format_type, output_dir):
        queryset = Event.objects.all()
        filename = os.path.join(output_dir, f'events.{format_type}')
        
        self.stdout.write(f'Exporting {queryset.count()} events to {filename}...')
        
        if format_type == 'csv':
            self._write_csv(queryset, filename, EVENT_FIELDS)
        elif format_type == 'excel':
            self._write_excel(queryset, filename, EVENT_FIELDS)
        elif format_type == 'pdf':
            self._write_pdf(queryset, filename, EVENT_FIELDS, 'Events Report')
        
        self.stdout.write(self.style.SUCCESS(f'Events exported to {filename}'))

    def export_registrations(self, format_type, output_dir, options):
        queryset = EventRegistration.objects.all()
        
        if options.get('event_id'):
            queryset = queryset.filter(event_id=options['event_id'])
        if options.get('status'):
            queryset = queryset.filter(approval_status=options['status'])
        
        filename = os.path.join(output_dir, f'registrations.{format_type}')
        
        self.stdout.write(f'Exporting {queryset.count()} registrations to {filename}...')
        
        if format_type == 'csv':
            self._write_csv(queryset, filename, REGISTRATION_FIELDS)
        elif format_type == 'excel':
            self._write_excel(queryset, filename, REGISTRATION_FIELDS)
        elif format_type == 'pdf':
            self._write_pdf(queryset, filename, REGISTRATION_FIELDS, 'Registrations Report')
        
        self.stdout.write(self.style.SUCCESS(f'Registrations exported to {filename}'))

    def export_approval_users(self, format_type, output_dir):
        queryset = ApprovalUser.objects.all()
        filename = os.path.join(output_dir, f'approval_users.{format_type}')
        
        self.stdout.write(f'Exporting {queryset.count()} approval users to {filename}...')
        
        if format_type == 'csv':
            self._write_csv(queryset, filename, APPROVAL_USER_FIELDS)
        elif format_type == 'excel':
            self._write_excel(queryset, filename, APPROVAL_USER_FIELDS)
        elif format_type == 'pdf':
            self._write_pdf(queryset, filename, APPROVAL_USER_FIELDS, 'Approval Users Report')
        
        self.stdout.write(self.style.SUCCESS(f'Approval users exported to {filename}'))

    def _write_csv(self, queryset, filename, fields):
        import csv
        with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            headers = [field['header'] for field in fields]
            writer.writerow(headers)
            
            for obj in queryset:
                row = []
                for field in fields:
                    value = ExportManager._get_field_value(obj, field['field'])
                    row.append(str(value) if value is not None else '')
                writer.writerow(row)

    def _write_excel(self, queryset, filename, fields):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Export"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = [field['header'] for field in fields]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = ExportManager._get_field_value(obj, field['field'])
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)

    def _write_pdf(self, queryset, filename, fields, title):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        doc = SimpleDocTemplate(filename, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        story = []
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # Prepare table data
        headers = [field['header'] for field in fields]
        data = [headers]
        
        for obj in queryset:
            row = []
            for field in fields:
                value = ExportManager._get_field_value(obj, field['field'])
                row.append(str(value) if value is not None else '')
            data.append(row)
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)