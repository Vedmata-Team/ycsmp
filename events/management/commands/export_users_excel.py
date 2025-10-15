from django.core.management.base import BaseCommand
from django.utils import timezone
from events.models import EventRegistration
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
from datetime import datetime

class Command(BaseCommand):
    help = 'Export all registered users to Excel with separate sheets for approved and non-approved users'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output-dir',
            type=str,
            default='exports',
            help='Output directory for the Excel file (default: exports)'
        )
        parser.add_argument(
            '--filename',
            type=str,
            help='Custom filename for the Excel file'
        )

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS('Starting user data export...'))
        
        # Create output directory if it doesn't exist
        output_dir = options['output_dir']
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.stdout.write(f'Created directory: {output_dir}')

        # Generate filename
        if options['filename']:
            filename = options['filename']
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
        else:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'user_registrations_{timestamp}.xlsx'
        
        filepath = os.path.join(output_dir, filename)

        try:
            # Get all registrations
            all_registrations = EventRegistration.objects.select_related(
                'event', 'responsibility', 'district_approver', 'upzone_approver', 
                'final_approver', 'rejected_by'
            ).all()

            self.stdout.write(f'Found {all_registrations.count()} total registrations')

            # Separate approved and non-approved users
            approved_users = all_registrations.filter(approval_status='approved')
            non_approved_users = all_registrations.exclude(approval_status='approved')

            self.stdout.write(f'Approved users: {approved_users.count()}')
            self.stdout.write(f'Non-approved users: {non_approved_users.count()}')

            # Create Excel workbook
            wb = Workbook()
            
            # Define headers
            headers = [
                'Registration Number', 'Full Name', 'Phone', 'Email', 'Date of Birth', 'Gender',
                'Registration Type', 'Event', 'State', 'District/City', 'Village/Taluka', 'Country',
                'Education', 'Occupation', 'Transport Mode', 'Vehicle Number', 'Arrival Date',
                'Previous Shivir', 'Gayatri Diksha', 'Special Skills', 'Selected Campaigns',
                'Selected Vibhags', 'Responsibility', 'Volunteer Start Date', 'Volunteer End Date',
                'Interested in Volunteering', 'Volunteering Details', 'Approval Status',
                'District Approver', 'District Approved At', 'UpZone Approver', 'UpZone Approved At',
                'Final Approver', 'Final Approved At', 'Rejected By', 'Rejected At', 'Rejection Reason',
                'Registration Date', 'Payment Status', 'Email Sent', 'Is Confirmed',
                'Aadhar Upload Type', 'Aadhar Full URL', 'Aadhar Front URL', 'Aadhar Back URL', 'Passport Photo URL'
            ]
            
            def get_registration_row(reg):
                return [
                    reg.registration_number or 'Not Generated',
                    reg.full_name,
                    reg.phone,
                    reg.email,
                    reg.date_of_birth.strftime('%d/%m/%Y') if reg.date_of_birth else '',
                    reg.get_gender_display(),
                    reg.get_registration_type_display(),
                    reg.event.title,
                    reg.state,
                    reg.city,
                    reg.village_taluka,
                    reg.country,
                    reg.get_education_display(),
                    reg.occupation or '',
                    reg.get_transport_mode_display(),
                    reg.vehicle_number or '',
                    reg.get_arrival_date_display(),
                    'Yes' if reg.previous_shivir else 'No',
                    'Yes' if reg.gayatri_diksha else ('No' if reg.gayatri_diksha is False else 'Not Specified'),
                    reg.special_skills_other or '',
                    reg.get_campaign_names(),
                    reg.get_vibhag_names(),
                    reg.responsibility.name if reg.responsibility else '',
                    reg.volunteer_start_date.strftime('%d/%m/%Y') if reg.volunteer_start_date else '',
                    reg.volunteer_end_date.strftime('%d/%m/%Y') if reg.volunteer_end_date else '',
                    'Yes' if reg.interested_in_volunteering else 'No',
                    reg.volunteering_details or '',
                    reg.get_approval_status_display(),
                    reg.district_approver.get_full_name() if reg.district_approver else '',
                    reg.district_approved_at.strftime('%d/%m/%Y %H:%M') if reg.district_approved_at else '',
                    reg.upzone_approver.get_full_name() if reg.upzone_approver else '',
                    reg.upzone_approved_at.strftime('%d/%m/%Y %H:%M') if reg.upzone_approved_at else '',
                    reg.final_approver.get_full_name() if reg.final_approver else '',
                    reg.final_approved_at.strftime('%d/%m/%Y %H:%M') if reg.final_approved_at else '',
                    reg.rejected_by.get_full_name() if reg.rejected_by else '',
                    reg.rejected_at.strftime('%d/%m/%Y %H:%M') if reg.rejected_at else '',
                    reg.rejection_reason or '',
                    reg.registration_date.strftime('%d/%m/%Y %H:%M'),
                    'Paid' if reg.payment_status else 'Pending',
                    'Yes' if reg.email_sent else 'No',
                    'Yes' if reg.is_confirmed else 'No',
                    reg.aadhar_upload_type or '',
                    reg.aadhar_full or '',
                    reg.aadhar_front or '',
                    reg.aadhar_back or '',
                    reg.passport_photo or '',
                ]
            
            # Create approved users sheet
            if approved_users.exists():
                ws_approved = wb.active
                ws_approved.title = "Approved Users"
                
                # Add headers with styling
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws_approved.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Add data
                for row_num, reg in enumerate(approved_users, 2):
                    row_data = get_registration_row(reg)
                    for col, value in enumerate(row_data, 1):
                        ws_approved.cell(row=row_num, column=col, value=value)
                
                self.stdout.write(f'✓ Approved users sheet created with {approved_users.count()} records')
            
            # Create non-approved users sheet
            if non_approved_users.exists():
                ws_non_approved = wb.create_sheet("Non-Approved Users")
                
                # Add headers with styling
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws_non_approved.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Add data
                for row_num, reg in enumerate(non_approved_users, 2):
                    row_data = get_registration_row(reg)
                    for col, value in enumerate(row_data, 1):
                        ws_non_approved.cell(row=row_num, column=col, value=value)
                
                self.stdout.write(f'✓ Non-approved users sheet created with {non_approved_users.count()} records')
            
            # Create summary sheet
            ws_summary = wb.create_sheet("Summary")
            summary_data = [
                ['Metric', 'Count'],
                ['Total Registrations', all_registrations.count()],
                ['Approved Users', approved_users.count()],
                ['Pending Users', all_registrations.filter(approval_status='pending').count()],
                ['District Approved Users', all_registrations.filter(approval_status='district_approved').count()],
                ['UpZone Approved Users', all_registrations.filter(approval_status='upzone_approved').count()],
                ['Rejected Users', all_registrations.filter(approval_status='rejected').count()],
                ['Participants', all_registrations.filter(registration_type='participant').count()],
                ['Volunteers', all_registrations.filter(registration_type='volunteer').count()],
                ['Organization Representatives', all_registrations.filter(registration_type='organization_representative').count()],
                ['Export Date', timezone.now().strftime('%d/%m/%Y %H:%M:%S')],
            ]
            
            for row_num, (metric, count) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_num, column=1, value=metric)
                ws_summary.cell(row=row_num, column=2, value=count)
                if row_num == 1:  # Header row
                    ws_summary.cell(row=row_num, column=1).font = Font(bold=True)
                    ws_summary.cell(row=row_num, column=2).font = Font(bold=True)
            
            self.stdout.write('✓ Summary sheet created')
            
            # Save the workbook
            wb.save(filepath)

            self.stdout.write(
                self.style.SUCCESS(f'✅ Export completed successfully!')
            )
            self.stdout.write(f'📁 File saved: {filepath}')
            self.stdout.write(f'📊 Total records exported: {all_registrations.count()}')

        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'❌ Export failed: {str(e)}')
            )
            raise e